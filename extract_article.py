import json
from openpyxl import Workbook
from datetime import datetime

with open('chatgpt/chat-gpt.json','rb') as json_file:
    json_data = json.load(json_file)

    # print('\n\nTitle')
    # print(json_data[0]['title'])
    # print('\n\nSnippet')
    # print(json_data[0]['snippet'])
    # print('\n\nBody')
    # print(json_data[0]['body'])
    # print('\n\nSource')
    # print(json_data[0]['source_name'])
    # print('\n\nPublication Date')
    # print(json_data[0]['publication_datetime'])
    # print('\n\nSubject Code')
    # print(json_data[0]['subject_codes'])
    # print('\n\nIndustry Code')
    # print(json_data[0]['industry_codes'])
    # print('\n\nCompany Code')
    # print(json_data[0]['company_codes'])
    #
    # print(len(json_data))

num_data = len(json_data)

to_extract = ['datanum', 'title', 'body','source_name','publication_datetime', 'num_paragraph' ]

wb = Workbook()
ws = wb.active

for k in range(len(to_extract)):
    ws.cell(row=1, column=(k + 1), value=to_extract[k])

for j in range(num_data):
    ws.cell(row=(j + 2), column=1, value=(j + 1))
    for k in range(1,len(to_extract)-1):
        if json_data[j][to_extract[k]]:
            if to_extract[k] == 'body':
                body_data = json_data[j]['snippet'] + "\n\n" + json_data[j]['body']
                ws.cell(row=(j + 2), column=(k + 1), value=body_data)
            else:
                if to_extract[k] == 'publication_datetime':
                    time = datetime.utcfromtimestamp(json_data[j][to_extract[k]]/1000).strftime("%Y%m%d")
                    ws.cell(row=(j + 2), column=(k + 1), value=time)
                else:
                    ws.cell(row=(j + 2), column=(k + 1), value=json_data[j][to_extract[k]])
        else:
            ws.cell(row=(j + 2), column=(k + 1), value="")
    num_paragraph = body_data.count("\n\n")+1
    ws.cell(row=(j + 2), column=(k + 2), value=num_paragraph)


wb.save("extracted_articles_0706.xlsx")
