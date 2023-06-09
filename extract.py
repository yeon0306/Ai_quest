import json
from datetime import datetime

with open('chat-gpt.json', 'r') as json_file:
    json_data = json.load(json_file)

# print test whether given data is loaded properly
print(len(json_data))
print("\n\nTitle")
print(json_data[0]['title'])
print("\n\nSnippet")
print(json_data[0]['snippet'])
print("\n\nBody")
print(json_data[0]['body'])
print("\n\nSource Name")
print(json_data[0]['source_name'])
print("\n\nPublication Datetime")
unix_time = json_data[0]['publication_datetime']
pub_time = datetime.utcfromtimestamp(unix_time/1000).strftime("%Y%m%d")
print(pub_time)

# body = snippet + body
to_extract = ['datanum', 'title', 'snippet', 'body', 'source_name', 'publication_datetime', 'num_paragraph']

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for j in range(len(to_extract)):
    ws.cell(row=1, column=(j+1), value=to_extract[j])

#num_data = len(json_data)
num_data = len(json_data)

for j in range(num_data):
    ws.cell(row=(j+2), column=1, value=(j+1))
    for k in range(1,len(to_extract)-1):
        key = to_extract[k]

        if key == 'publication_datetime':
            unix_time = json_data[j][key]
            pub_time = datetime.utcfromtimestamp(unix_time / 1000).strftime("%Y%m%d")
            ws.cell(row=(j + 2), column=(k + 1), value=pub_time)
        else:
            ws.cell(row=(j + 2), column=(k + 1), value=json_data[j].get(key, ""))

        # 본문에서 단락 수를 계산합니다.
    snippet = json_data[j].get('snippet', "")
    body = json_data[j].get('body', "")

    if snippet is None:
        snippet = ""

    if body is None:
        body = ""

    num_paragraph = (snippet + body).count('\n') + 1
    ws.cell(row=(j + 2), column=len(to_extract), value=num_paragraph)

wb.save("test1.xlsx")